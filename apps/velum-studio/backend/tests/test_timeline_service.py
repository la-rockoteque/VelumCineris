from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.services.timeline_service import (
    TimelineService,
    _apply_holiday_updates_rows,
    _extract_present_holidays,
    _holiday_header_indexes,
    _merge_holidays,
    _normalize_holiday_updates,
    _parse_holidays,
)


def _create_timeline_workbook(path: Path) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    era = wb.create_sheet("Era")
    era["A1"] = "ignored"
    era["A2"] = "Year of the Ember"
    era["A3"] = 100
    era["A4"] = "Era event"

    calendar = wb.create_sheet("Calendar")
    calendar.append([None, "Month Name", "Description", "Chore Name", "Chore Description", "Deity Name", "Domain"])
    calendar.append(["1st Month", "FIRSTCOLD", "Deep winter", "Smith Day", "Forge tools", "Heraclus", "War"])

    naming = wb.create_sheet("Naming")
    naming.append(["A (Modifier)", "A (Phenomenon)", "A (Anchor)", "A (Modifier)"])
    naming.append(["Iron", "Dawn", "Citadel", "Red"])

    holidays = wb.create_sheet("Holidays")
    holidays["A1"] = "Fest of the Fools"

    present = wb.create_sheet("Present")
    present.append(["Year of Ember", None, None, None, None])
    present.append(["FIRSTCOLD", None, None, None, None])
    present.append(["Avenoir", "Exulansis", "Ambedo", "Opia", "Liberosis"])
    present.append([1, 2, 3, 4, 5])
    present.append(["Dies Natalis", None, None, None, None])
    present.append([6, 7, 8, 9, 10])
    present.append(["Equinox", None, None, None, None])

    wb.save(path)
    wb.close()


def test_holiday_header_indexes_supports_extended_columns() -> None:
    headers = ["Name", "Month", "Day", "Recurrence", "Source", "Weekday", "Year", "Notes"]
    mapping = _holiday_header_indexes(headers)

    assert mapping["name"] == 0
    assert mapping["month_name"] == 1
    assert mapping["day"] == 2
    assert mapping["recurrence"] == 3
    assert mapping["source"] == 4
    assert mapping["weekday"] == 5
    assert mapping["year"] == 6
    assert mapping["notes"] == 7


def test_parse_holidays_reads_extended_metadata_columns() -> None:
    rows = [
        ["Name", "Month", "Day", "Recurrence", "Source", "Weekday", "Year", "Notes"],
        ["Festival of Lanterns", "FIRSTCOLD", 12, "yearly", "holidays", "Avenoir", "100", "City-wide lights"],
    ]

    parsed = _parse_holidays(rows)
    assert len(parsed) == 1
    holiday = parsed[0]
    assert holiday["name"] == "Festival of Lanterns"
    assert holiday["month_name"] == "FIRSTCOLD"
    assert holiday["day"] == 12
    assert holiday["recurrence"] == "yearly"
    assert holiday["source"] == "holidays"
    assert holiday["weekday"] == "Avenoir"
    assert holiday["year"] == "100"
    assert holiday["notes"] == "City-wide lights"


def test_parse_holidays_without_headers_uses_fallback_positions() -> None:
    rows = [["Founders Day", "FIRSTCOLD", 3, "yearly", "manual", "Avenoir", "100", "Core lore day"]]
    parsed = _parse_holidays(rows)

    assert len(parsed) == 1
    holiday = parsed[0]
    assert holiday["name"] == "Founders Day"
    assert holiday["month_name"] == "FIRSTCOLD"
    assert holiday["day"] == 3
    assert holiday["recurrence"] == "yearly"
    assert holiday["source"] == "manual"
    assert holiday["weekday"] == "Avenoir"
    assert holiday["year"] == "100"
    assert holiday["notes"] == "Core lore day"


def test_extract_present_holidays_includes_weekday_and_dedupes() -> None:
    present_months = [
        {
            "month_name": "FIRSTCOLD",
            "year_name": "Year 100",
            "weeks": [
                {
                    "week_index": 1,
                    "days": [
                        {"weekday": "Avenoir", "day": 1, "event": "Dies Natalis"},
                        {"weekday": "Exulansis", "day": 2, "event": "Dies Natalis"},
                    ],
                },
                {
                    "week_index": 2,
                    "days": [
                        {"weekday": "Avenoir", "day": 1, "event": "Dies Natalis"},
                    ],
                },
            ],
        }
    ]

    parsed = _extract_present_holidays(present_months)
    assert len(parsed) == 2
    assert parsed[0]["name"] == "Dies Natalis"
    assert parsed[0]["weekday"] == "Avenoir"
    assert parsed[0]["year"] == "Year 100"


def test_merge_holidays_dedupes_by_name_month_day() -> None:
    base = [{"name": "Dies Natalis", "month_name": "FIRSTCOLD", "day": 1, "source": "holidays"}]
    present = [{"name": "Dies Natalis", "month_name": "FIRSTCOLD", "day": 1, "source": "present"}]

    merged = _merge_holidays(base, present)
    assert len(merged) == 1
    assert merged[0]["source"] == "holidays"


def test_normalize_holiday_updates_dedupes_and_adds_defaults() -> None:
    updates = [
        {"name": "Equinox", "month_name": "FIRSTCOLD", "day": 6},
        {"name": "Equinox", "month_name": "FIRSTCOLD", "day": 6},
    ]

    normalized = _normalize_holiday_updates(updates)
    assert len(normalized) == 1
    assert normalized[0]["recurrence"] == "yearly"
    assert normalized[0]["source"] == "holidays"
    assert normalized[0]["weekday"] is None
    assert normalized[0]["year"] is None
    assert normalized[0]["notes"] is None


def test_apply_holiday_updates_rows_writes_full_header_set() -> None:
    rows = [["Name"], ["Legacy"]]
    updated = _apply_holiday_updates_rows(
        rows,
        [
            {
                "name": "Dies Natalis",
                "month_name": "FIRSTCOLD",
                "day": 1,
                "recurrence": "yearly",
                "source": "present",
                "weekday": "Avenoir",
                "year": "Year 100",
                "notes": "Imported",
            }
        ],
    )

    assert updated[0] == ["Name", "Month", "Day", "Recurrence", "Source", "Weekday", "Year", "Notes"]
    assert updated[1] == ["Dies Natalis", "FIRSTCOLD", 1, "yearly", "present", "Avenoir", "Year 100", "Imported"]


def test_load_catalog_syncs_present_events_into_holidays_sheet(tmp_path: Path) -> None:
    workbook_path = tmp_path / "timeline.xlsx"
    _create_timeline_workbook(workbook_path)

    service = TimelineService(
        timeline_xlsx_path=workbook_path,
        timeline_spreadsheet_id="unused",
        credentials_path=tmp_path / "missing-key.json",
    )
    payload = service.load_catalog("xlsx")

    assert payload["summary"]["present_months"] >= 1
    assert payload["summary"]["holidays"] >= 3

    wb = load_workbook(workbook_path, data_only=True)
    try:
        ws = wb["Holidays"]
        headers = [ws.cell(1, col).value for col in range(1, 9)]
        assert headers == ["Name", "Month", "Day", "Recurrence", "Source", "Weekday", "Year", "Notes"]

        names = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]
        assert "Fest of the Fools" in names
        assert "Dies Natalis" in names
        assert "Equinox" in names
    finally:
        wb.close()


def test_load_catalog_holiday_sync_is_idempotent(tmp_path: Path) -> None:
    workbook_path = tmp_path / "timeline.xlsx"
    _create_timeline_workbook(workbook_path)

    service = TimelineService(
        timeline_xlsx_path=workbook_path,
        timeline_spreadsheet_id="unused",
        credentials_path=tmp_path / "missing-key.json",
    )
    service.load_catalog("xlsx")

    wb = load_workbook(workbook_path, data_only=True)
    try:
        before_rows = wb["Holidays"].max_row
    finally:
        wb.close()

    service.load_catalog("xlsx")

    wb = load_workbook(workbook_path, data_only=True)
    try:
        after_rows = wb["Holidays"].max_row
    finally:
        wb.close()

    assert after_rows == before_rows
