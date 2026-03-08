from __future__ import annotations

from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import column_index_from_string

from Spreadsheet.core.Helpers.sheets import SpreadsheetClient


VALID_SOURCES = {"auto", "google", "xlsx"}
HOLIDAY_HEADERS = ["Name", "Month", "Day", "Recurrence", "Source", "Weekday", "Year", "Notes"]


class TimelineService:
    def __init__(
        self,
        *,
        timeline_xlsx_path: Path,
        timeline_spreadsheet_id: str,
        credentials_path: Path,
    ) -> None:
        self.timeline_xlsx_path = timeline_xlsx_path
        self.timeline_spreadsheet_id = timeline_spreadsheet_id
        self.credentials_path = credentials_path
        self._google_client: SpreadsheetClient | None = None

    def load_catalog(self, source: str) -> dict[str, Any]:
        resolved = self._resolve_source(source)
        sheets = self._read_sheets(
            resolved,
            ("Era", "Calendar", "Naming", "Holidays", "Present"),
        )

        calendar_months = _parse_calendar_months(sheets.get("Calendar", []))
        naming_groups = _parse_naming_groups(sheets.get("Naming", []))
        present_months = _parse_present_months(sheets.get("Present", []), calendar_months)
        weekdays = present_months[0]["weekdays"] if present_months else []
        holidays = _merge_holidays(
            _parse_holidays(sheets.get("Holidays", [])),
            _extract_present_holidays(present_months),
        )
        if resolved == "xlsx":
            _sync_holidays_sheet_xlsx(self.timeline_xlsx_path, holidays)
        era_events = _parse_era_events(sheets.get("Era", []))
        naming_template = _derive_naming_template(naming_groups)

        return {
            "source": resolved,
            "calendar_months": calendar_months,
            "naming_groups": naming_groups,
            "naming_template": naming_template,
            "weekdays": weekdays,
            "holidays": holidays,
            "era_events": era_events,
            "present_months": present_months,
            "summary": {
                "months": len(calendar_months),
                "naming_groups": len(naming_groups),
                "weekdays": len(weekdays),
                "holidays": len(holidays),
                "era_events": len(era_events),
                "present_months": len(present_months),
            },
        }

    def save_catalog(
        self,
        source: str,
        *,
        calendar_months: list[dict[str, Any]] | None = None,
        naming_groups: list[dict[str, Any]] | None = None,
        weekdays: list[str] | None = None,
        holidays: list[dict[str, Any]] | None = None,
        era_events: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any]:
        resolved = self._resolve_source(source)
        if resolved == "xlsx":
            self._save_xlsx(
                calendar_months=calendar_months,
                naming_groups=naming_groups,
                weekdays=weekdays,
                holidays=holidays,
                era_events=era_events,
            )
        elif resolved == "google":
            self._save_google(
                calendar_months=calendar_months,
                naming_groups=naming_groups,
                weekdays=weekdays,
                holidays=holidays,
                era_events=era_events,
            )
        else:
            raise ValueError("No available timeline source.")

        return self.load_catalog(source)

    def _save_xlsx(
        self,
        *,
        calendar_months: list[dict[str, Any]] | None,
        naming_groups: list[dict[str, Any]] | None,
        weekdays: list[str] | None,
        holidays: list[dict[str, Any]] | None,
        era_events: list[dict[str, Any]] | None,
    ) -> None:
        if not self.timeline_xlsx_path.exists():
            raise FileNotFoundError(f"Timeline workbook not found: {self.timeline_xlsx_path}")

        wb = load_workbook(self.timeline_xlsx_path)
        try:
            if calendar_months is not None and "Calendar" in wb.sheetnames:
                _apply_calendar_updates_xlsx(wb["Calendar"], calendar_months)
            if naming_groups is not None and "Naming" in wb.sheetnames:
                _apply_naming_updates_xlsx(wb["Naming"], naming_groups)
            if weekdays is not None and "Present" in wb.sheetnames:
                _apply_weekday_updates_xlsx(wb["Present"], weekdays)
            if holidays is not None and "Holidays" in wb.sheetnames:
                _apply_holiday_updates_xlsx(wb["Holidays"], holidays)
            if era_events is not None and "Era" in wb.sheetnames:
                _apply_era_updates_xlsx(wb["Era"], era_events)

            wb.save(self.timeline_xlsx_path)
        finally:
            wb.close()

    def _save_google(
        self,
        *,
        calendar_months: list[dict[str, Any]] | None,
        naming_groups: list[dict[str, Any]] | None,
        weekdays: list[str] | None,
        holidays: list[dict[str, Any]] | None,
        era_events: list[dict[str, Any]] | None,
    ) -> None:
        client = self._get_google_client()

        if calendar_months is not None:
            rows = client.get_rows_by_title("Calendar")
            rows = _apply_calendar_updates_rows(rows, calendar_months)
            _write_sheet_rows_google(client, "Calendar", rows)

        if naming_groups is not None:
            rows = client.get_rows_by_title("Naming")
            rows = _apply_naming_updates_rows(rows, naming_groups)
            _write_sheet_rows_google(client, "Naming", rows)

        if weekdays is not None:
            rows = client.get_rows_by_title("Present")
            rows = _apply_weekday_updates_rows(rows, weekdays)
            _write_sheet_rows_google(client, "Present", rows)

        if holidays is not None:
            rows = client.get_rows_by_title("Holidays")
            rows = _apply_holiday_updates_rows(rows, holidays)
            _write_sheet_rows_google(client, "Holidays", rows)

        if era_events is not None:
            rows = client.get_rows_by_title("Era")
            rows = _apply_era_updates_rows(rows, era_events)
            _write_sheet_rows_google(client, "Era", rows)

    def _read_sheets(self, source: str, sheet_names: tuple[str, ...]) -> dict[str, list[list[Any]]]:
        if source == "xlsx":
            if not self.timeline_xlsx_path.exists():
                raise FileNotFoundError(f"Timeline workbook not found: {self.timeline_xlsx_path}")
            wb = load_workbook(self.timeline_xlsx_path, read_only=True, data_only=False)
            try:
                data: dict[str, list[list[Any]]] = {}
                for name in sheet_names:
                    if name not in wb.sheetnames:
                        data[name] = []
                        continue
                    ws = wb[name]
                    data[name] = [list(row) for row in ws.iter_rows(values_only=True)]
                return data
            finally:
                wb.close()

        if source == "google":
            client = self._get_google_client()
            existing = set(client.list_sheet_names())
            data: dict[str, list[list[Any]]] = {}
            for name in sheet_names:
                if name not in existing:
                    data[name] = []
                    continue
                data[name] = client.get_rows_by_title(name)
            return data

        raise ValueError("Invalid timeline source.")

    def _resolve_source(self, source: str) -> str:
        normalized = source.strip().lower()
        if normalized not in VALID_SOURCES:
            raise ValueError("Invalid source. Use auto, xlsx, or google.")

        if normalized == "auto":
            if self.timeline_xlsx_path.exists():
                return "xlsx"
            if self.credentials_path.exists():
                return "google"
            raise ValueError("No available timeline source in auto mode.")

        if normalized == "xlsx":
            if not self.timeline_xlsx_path.exists():
                raise FileNotFoundError(f"Timeline workbook not found: {self.timeline_xlsx_path}")
            return "xlsx"

        if not self.credentials_path.exists():
            raise FileNotFoundError(f"Google credentials not found: {self.credentials_path}")
        return "google"

    def _get_google_client(self) -> SpreadsheetClient:
        if self._google_client is None:
            self._google_client = SpreadsheetClient(
                self.timeline_spreadsheet_id,
                credentials_path=self.credentials_path,
                scopes=(
                    "https://www.googleapis.com/auth/spreadsheets",
                    "https://www.googleapis.com/auth/drive",
                ),
            )
        return self._google_client


def _parse_calendar_months(rows: list[list[Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []

    header_row_idx = _detect_header_row_index(rows)
    headers = rows[header_row_idx] if header_row_idx < len(rows) else []
    header_map = _calendar_header_indexes(headers)

    out: list[dict[str, Any]] = []
    for idx in range(header_row_idx + 1, len(rows)):
        row = rows[idx]
        if not any(_as_text(cell) for cell in row):
            continue

        out.append(
            {
                "row_number": idx + 1,
                "month_order": _cell_at(row, header_map.get("month_order", 0)),
                "month_name": _cell_at(row, header_map.get("month_name", 1)),
                "description": _cell_at(row, header_map.get("description", 2)),
                "chore_name": _cell_at(row, header_map.get("chore_name", 3)),
                "chore_description": _cell_at(row, header_map.get("chore_description", 4)),
                "deity_name": _cell_at(row, header_map.get("deity_name", 5)),
                "domain": _cell_at(row, header_map.get("domain", 6)),
            }
        )
    return out


def _parse_naming_groups(rows: list[list[Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []

    headers = rows[0] if rows else []
    seen_labels: dict[str, int] = {}
    groups: list[dict[str, Any]] = []

    for col_idx, raw_header in enumerate(headers, start=1):
        label = _as_text(raw_header)
        if not label:
            continue

        count = seen_labels.get(label, 0) + 1
        seen_labels[label] = count
        display = label if count == 1 else f"{label} ({count})"

        values: list[str] = []
        for row in rows[1:]:
            value = _as_text(_cell_at(row, col_idx - 1))
            if value:
                values.append(value)

        groups.append(
            {
                "key": f"col_{col_idx}",
                "label": display,
                "values": values,
            }
        )

    return groups


def _derive_naming_template(groups: list[dict[str, Any]]) -> str:
    if not groups:
        return "Year of the {A (Modifier)} {A (Phenomenon)} Century of the {A (Anchor)} {A (Modifier)}"

    primary = groups[0]["label"] if len(groups) >= 1 else "A (Modifier)"
    phenomenon = groups[1]["label"] if len(groups) >= 2 else "A (Phenomenon)"
    anchor = groups[2]["label"] if len(groups) >= 3 else "A (Anchor)"
    trailing = groups[3]["label"] if len(groups) >= 4 else "A (Modifier)"

    def strip_variant(label: str) -> str:
        return re.sub(r"\s+\(\d+\)$", "", str(label or "")).strip() or "Part"

    return (
        f"Year of the {{{strip_variant(primary)}}} {{{strip_variant(phenomenon)}}} "
        f"Century of the {{{strip_variant(anchor)}}} {{{strip_variant(trailing)}}}"
    )


def _parse_holidays(rows: list[list[Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []

    header_idx = 0
    header_map = _holiday_header_indexes(rows[0])
    if "name" not in header_map:
        header_map = {}

    start_row = 2 if header_map else 1
    holidays: list[dict[str, Any]] = []

    for row_number in range(start_row, len(rows) + 1):
        row = rows[row_number - 1]
        if not any(_as_text(cell) for cell in row):
            continue

        if header_map:
            name = _as_text(_cell_at(row, header_map.get("name", 0)))
            month_name = _as_text(_cell_at(row, header_map.get("month_name", -1))) if "month_name" in header_map else ""
            day = _to_int(_cell_at(row, header_map.get("day", -1))) if "day" in header_map else None
            recurrence = _as_text(_cell_at(row, header_map.get("recurrence", -1))) if "recurrence" in header_map else ""
            source = _as_text(_cell_at(row, header_map.get("source", -1))) if "source" in header_map else ""
            weekday = _as_text(_cell_at(row, header_map.get("weekday", -1))) if "weekday" in header_map else ""
            year = _as_text(_cell_at(row, header_map.get("year", -1))) if "year" in header_map else ""
            notes = _as_text(_cell_at(row, header_map.get("notes", -1))) if "notes" in header_map else ""
        else:
            name = _as_text(_cell_at(row, 0))
            month_name = _as_text(_cell_at(row, 1))
            day = _to_int(_cell_at(row, 2))
            recurrence = _as_text(_cell_at(row, 3))
            source = _as_text(_cell_at(row, 4))
            weekday = _as_text(_cell_at(row, 5))
            year = _as_text(_cell_at(row, 6))
            notes = _as_text(_cell_at(row, 7))

        if not name:
            continue
        holidays.append(
            {
                "row_number": row_number,
                "name": name,
                "month_name": month_name or None,
                "day": day,
                "recurrence": (recurrence or ("yearly" if month_name or day else None)),
                "source": source or "holidays",
                "weekday": weekday or None,
                "year": year or None,
                "notes": notes or None,
            }
        )
    return holidays


def _extract_present_holidays(present_months: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    seen: set[tuple[str, str, int | None]] = set()

    for month in present_months:
        month_name = _as_text(month.get("month_name"))
        weeks = month.get("weeks", [])
        if not isinstance(weeks, list):
            continue

        for week in weeks:
            days = week.get("days", []) if isinstance(week, dict) else []
            if not isinstance(days, list):
                continue
            for day_payload in days:
                if not isinstance(day_payload, dict):
                    continue
                day_number = _to_int(day_payload.get("day"))
                event_text = _as_text(day_payload.get("event"))
                if not event_text:
                    continue

                fragments = re.split(r"[\n/]+", event_text)
                for fragment in fragments:
                    name = fragment.strip()
                    if not name:
                        continue
                    key = (name.lower(), month_name.lower(), day_number)
                    if key in seen:
                        continue
                    seen.add(key)
                    out.append(
                        {
                            "row_number": None,
                            "name": name,
                            "month_name": month_name or None,
                            "day": day_number,
                            "recurrence": "yearly",
                            "source": "present",
                            "weekday": _as_text(day_payload.get("weekday")) or None,
                            "year": _as_text(month.get("year_name")) or None,
                            "notes": None,
                        }
                    )
    return out


def _merge_holidays(base: list[dict[str, Any]], present: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: list[dict[str, Any]] = []
    seen: set[tuple[str, str, int | None]] = set()

    for item in base + present:
        name = _as_text(item.get("name"))
        month_name = _as_text(item.get("month_name"))
        day = _to_int(item.get("day"))
        if not name:
            continue
        key = (name.lower(), month_name.lower(), day)
        if key in seen:
            continue
        seen.add(key)
        payload = dict(item)
        payload["month_name"] = month_name or None
        payload["day"] = day
        payload["recurrence"] = _as_text(item.get("recurrence")) or ("yearly" if month_name or day else None)
        payload["source"] = _as_text(item.get("source")) or "holidays"
        merged.append(payload)

    return merged


def _holiday_header_indexes(headers: list[Any]) -> dict[str, int]:
    normalized = {re.sub(r"[^a-z0-9]+", "", _as_text(value).lower()): idx for idx, value in enumerate(headers)}
    mapping: dict[str, int] = {}

    if "name" in normalized:
        mapping["name"] = normalized["name"]
    elif "holiday" in normalized:
        mapping["name"] = normalized["holiday"]

    if "month" in normalized:
        mapping["month_name"] = normalized["month"]
    elif "monthname" in normalized:
        mapping["month_name"] = normalized["monthname"]

    if "day" in normalized:
        mapping["day"] = normalized["day"]
    elif "dayofmonth" in normalized:
        mapping["day"] = normalized["dayofmonth"]

    if "recurrence" in normalized:
        mapping["recurrence"] = normalized["recurrence"]
    elif "repeat" in normalized:
        mapping["recurrence"] = normalized["repeat"]

    if "source" in normalized:
        mapping["source"] = normalized["source"]
    if "weekday" in normalized:
        mapping["weekday"] = normalized["weekday"]
    elif "dayname" in normalized:
        mapping["weekday"] = normalized["dayname"]
    if "year" in normalized:
        mapping["year"] = normalized["year"]
    if "notes" in normalized:
        mapping["notes"] = normalized["notes"]
    elif "note" in normalized:
        mapping["notes"] = normalized["note"]

    return mapping


def _extract_weekdays(rows: list[list[Any]]) -> tuple[list[str], list[int]]:
    present = _parse_present_months(rows)
    if not present:
        return [], []

    weekdays = present[0].get("weekdays", [])
    hit_rows = [item.get("row_start", 0) for item in present if item.get("row_start")]
    return weekdays, hit_rows


def _parse_era_events(rows: list[list[Any]], limit: int = 1200) -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []
    if not rows:
        return events

    # Era format:
    # row 1 ignored
    # row 2 names, row 3 year numbers, rows 4-7 events
    # row 8 ignored, then pattern repeats from row 9
    block_start = 1  # zero-based, corresponds to row 2 in sheet
    resolved_year_cells: dict[tuple[int, int], int] = {}
    while block_start < len(rows):
        name_row = rows[block_start] if block_start < len(rows) else []
        year_row = rows[block_start + 1] if block_start + 1 < len(rows) else []
        event_rows = [
            rows[block_start + 2] if block_start + 2 < len(rows) else [],
            rows[block_start + 3] if block_start + 3 < len(rows) else [],
            rows[block_start + 4] if block_start + 4 < len(rows) else [],
            rows[block_start + 5] if block_start + 5 < len(rows) else [],
        ]

        width = max(len(name_row), len(year_row), *(len(row) for row in event_rows), 0)
        if width == 0:
            block_start += 7
            continue

        any_year = False
        previous_year: int | None = None
        year_row_number = block_start + 2
        for col_idx in range(width):
            raw_year = _cell_at(year_row, col_idx)
            year_value = _to_int(raw_year)
            if year_value is None:
                year_value = _infer_formula_number(
                    raw_year,
                    previous_year,
                    resolved_year_cells,
                )
            if year_value is None:
                continue
            previous_year = year_value
            resolved_year_cells[(year_row_number, col_idx + 1)] = year_value
            any_year = True
            year_name = _as_text(_cell_at(name_row, col_idx))
            era_name = _extract_era_name(year_name)

            for offset, event_row in enumerate(event_rows):
                event_text = _as_text(_cell_at(event_row, col_idx))
                if not event_text:
                    continue
                events.append(
                    {
                        "year": year_value,
                        "era": era_name or None,
                        "event": event_text,
                        "row_number": block_start + 3 + offset,
                        "column": col_idx + 1,
                    }
                )
                if len(events) >= limit:
                    break
            if len(events) >= limit:
                break
        if len(events) >= limit:
            break

        # If no years were detected in this block, stop if remaining rows are empty.
        if not any_year:
            trailing_has_data = any(
                _as_text(cell)
                for row in rows[block_start : min(len(rows), block_start + 7)]
                for cell in row
            )
            if not trailing_has_data:
                break

        block_start += 7

    events.sort(key=lambda item: (item["year"] if item["year"] is not None else 10**12, item["row_number"], item["column"]))
    return events


def _parse_present_months(
    rows: list[list[Any]],
    calendar_months: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    months: list[dict[str, Any]] = []
    if not rows:
        return months

    max_cols = max((len(row) for row in rows), default=0)
    if max_cols == 0:
        return months

    row_idx = 0
    while row_idx + 2 < len(rows):
        year_row = rows[row_idx]
        month_row = rows[row_idx + 1]
        weekday_row = rows[row_idx + 2]

        group_found = False
        col = 0
        while col < max_cols:
            year_name = _as_text(_cell_at(year_row, col))
            if year_name.startswith("=") or "ArrayFormula" in year_name:
                year_name = ""
            month_name = _as_text(_cell_at(month_row, col))
            if month_name.startswith("=") and calendar_months:
                idx = len(months) % max(len(calendar_months), 1)
                month_name = _as_text(calendar_months[idx].get("month_name")).upper() or month_name
            weekdays = [_as_text(_cell_at(weekday_row, col + offset)) for offset in range(5)]
            weekdays = [value for value in weekdays if value]

            # Merged cells keep value only in first column; weekdays occupy 5 columns.
            if month_name and len(weekdays) >= 3:
                group_found = True
                month_payload = {
                    "row_start": row_idx + 1,
                    "column_start": col + 1,
                    "year_name": year_name,
                    "month_name": month_name,
                    "weekdays": weekdays[:5],
                    "weeks": [],
                }

                last_day_number: int | None = None
                for week_idx in range(4):
                    number_row_idx = row_idx + 3 + (week_idx * 2)
                    events_row_idx = number_row_idx + 1
                    if number_row_idx >= len(rows):
                        break
                    number_row = rows[number_row_idx]
                    events_row = rows[events_row_idx] if events_row_idx < len(rows) else []

                    days: list[dict[str, Any]] = []
                    has_any = False
                    for day_offset in range(5):
                        day_col = col + day_offset
                        raw_day_number = _cell_at(number_row, day_col)
                        day_number = _to_int(raw_day_number)
                        if day_number is None:
                            day_number = _infer_formula_number(raw_day_number, last_day_number)
                        day_event = _as_text(_cell_at(events_row, day_col))
                        if day_number is not None or day_event:
                            has_any = True
                        if day_number is not None:
                            last_day_number = day_number
                        days.append(
                            {
                                "weekday": month_payload["weekdays"][day_offset]
                                if day_offset < len(month_payload["weekdays"])
                                else "",
                                "day": day_number,
                                "event": day_event or None,
                            }
                        )
                    if has_any:
                        month_payload["weeks"].append({"week_index": week_idx + 1, "days": days})

                month_payload["day_count"] = sum(
                    1
                    for week in month_payload["weeks"]
                    for day in week["days"]
                    if day.get("day") is not None
                )
                months.append(month_payload)
                col += 6  # 5 days + 1 spacer column
                continue

            col += 1

        if group_found:
            row_idx += 12  # next set of 4 months starts lower
        else:
            row_idx += 1

    return months


def _extract_era_name(value: str) -> str:
    if not value:
        return ""
    if "ArrayFormula" in value:
        return ""
    parts = [part.strip() for part in value.splitlines() if part.strip()]
    if len(parts) >= 2:
        return parts[1]
    match = re.search(r"(Century of [^\n]+)", value, flags=re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return parts[0] if parts else ""


def _apply_calendar_updates_xlsx(ws: Any, updates: list[dict[str, Any]]) -> None:
    header_map = _calendar_header_indexes([ws.cell(1, col).value for col in range(1, ws.max_column + 1)])
    for item in updates:
        row_number = _to_int(item.get("row_number"))
        if row_number is None or row_number < 2:
            continue
        ws.cell(row_number, header_map.get("month_order", 0) + 1).value = _clean_for_write(item.get("month_order"))
        ws.cell(row_number, header_map.get("month_name", 1) + 1).value = _clean_for_write(item.get("month_name"))
        ws.cell(row_number, header_map.get("description", 2) + 1).value = _clean_for_write(item.get("description"))
        ws.cell(row_number, header_map.get("chore_name", 3) + 1).value = _clean_for_write(item.get("chore_name"))
        ws.cell(row_number, header_map.get("chore_description", 4) + 1).value = _clean_for_write(
            item.get("chore_description")
        )
        ws.cell(row_number, header_map.get("deity_name", 5) + 1).value = _clean_for_write(item.get("deity_name"))
        ws.cell(row_number, header_map.get("domain", 6) + 1).value = _clean_for_write(item.get("domain"))


def _apply_naming_updates_xlsx(ws: Any, groups: list[dict[str, Any]]) -> None:
    for group in groups:
        col_idx = _column_from_group_key(group.get("key"))
        if col_idx is None:
            continue
        values = [str(value).strip() for value in group.get("values", []) if str(value).strip()]
        for row in range(2, max(ws.max_row, len(values) + 1) + 1):
            ws.cell(row, col_idx).value = None
        for offset, value in enumerate(values, start=2):
            ws.cell(offset, col_idx).value = value


def _apply_weekday_updates_xlsx(ws: Any, weekdays: list[str]) -> None:
    rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(1, ws.max_row + 1)]
    present_months = _parse_present_months(rows)
    if not present_months:
        return

    canonical = [str(value).strip() for value in weekdays if str(value).strip()]
    if not canonical:
        return

    for month in present_months:
        row_number = _to_int(month.get("row_start"))
        col_number = _to_int(month.get("column_start"))
        existing = month.get("weekdays", [])
        if row_number is None or col_number is None or not existing:
            continue
        for idx, current in enumerate(existing[:5]):
            next_value = canonical[idx] if idx < len(canonical) else str(current).strip()
            if next_value:
                ws.cell(row_number + 2, col_number + idx).value = next_value


def _apply_holiday_updates_xlsx(ws: Any, holidays: list[dict[str, Any]]) -> None:
    cleaned = _normalize_holiday_updates(holidays)
    headers = HOLIDAY_HEADERS

    target_rows = max(ws.max_row, len(cleaned) + 1)
    target_cols = max(ws.max_column, len(headers))

    for row in range(1, target_rows + 1):
        for col in range(1, target_cols + 1):
            ws.cell(row, col).value = None

    for col, value in enumerate(headers, start=1):
        ws.cell(1, col).value = value

    for row_idx, item in enumerate(cleaned, start=2):
        ws.cell(row_idx, 1).value = item.get("name")
        ws.cell(row_idx, 2).value = item.get("month_name")
        ws.cell(row_idx, 3).value = item.get("day")
        ws.cell(row_idx, 4).value = item.get("recurrence")
        ws.cell(row_idx, 5).value = item.get("source")
        ws.cell(row_idx, 6).value = item.get("weekday")
        ws.cell(row_idx, 7).value = item.get("year")
        ws.cell(row_idx, 8).value = item.get("notes")


def _apply_era_updates_xlsx(ws: Any, era_events: list[dict[str, Any]]) -> None:
    rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(1, ws.max_row + 1)]
    clear_rows, placements = _plan_era_event_placements(rows, era_events)
    if not clear_rows and not placements:
        return

    for row_number in clear_rows:
        for col in range(1, ws.max_column + 1):
            _set_cell_if_writable(ws, row_number, col, None)

    for row_number, col_number, text in placements:
        _set_cell_if_writable(ws, row_number, col_number, text)


def _apply_calendar_updates_rows(rows: list[list[Any]], updates: list[dict[str, Any]]) -> list[list[Any]]:
    matrix = _pad_rows(rows)
    if not matrix:
        return matrix

    header_map = _calendar_header_indexes(matrix[0])
    for item in updates:
        row_number = _to_int(item.get("row_number"))
        if row_number is None or row_number < 2:
            continue
        _ensure_row(matrix, row_number)
        row = matrix[row_number - 1]
        row[header_map.get("month_order", 0)] = _clean_for_write(item.get("month_order"))
        row[header_map.get("month_name", 1)] = _clean_for_write(item.get("month_name"))
        row[header_map.get("description", 2)] = _clean_for_write(item.get("description"))
        row[header_map.get("chore_name", 3)] = _clean_for_write(item.get("chore_name"))
        row[header_map.get("chore_description", 4)] = _clean_for_write(item.get("chore_description"))
        row[header_map.get("deity_name", 5)] = _clean_for_write(item.get("deity_name"))
        row[header_map.get("domain", 6)] = _clean_for_write(item.get("domain"))
    return matrix


def _apply_naming_updates_rows(rows: list[list[Any]], groups: list[dict[str, Any]]) -> list[list[Any]]:
    matrix = _pad_rows(rows)
    if not matrix:
        matrix = [[]]

    for group in groups:
        col_idx = _column_from_group_key(group.get("key"))
        if col_idx is None:
            continue

        values = [str(value).strip() for value in group.get("values", []) if str(value).strip()]
        for row in range(2, max(len(matrix), len(values) + 1) + 1):
            _ensure_row(matrix, row)
            _ensure_col(matrix[row - 1], col_idx)
            matrix[row - 1][col_idx - 1] = ""

        for row, value in enumerate(values, start=2):
            _ensure_row(matrix, row)
            _ensure_col(matrix[row - 1], col_idx)
            matrix[row - 1][col_idx - 1] = value
    return matrix


def _apply_weekday_updates_rows(rows: list[list[Any]], weekdays: list[str]) -> list[list[Any]]:
    matrix = _pad_rows(rows)
    present_months = _parse_present_months(matrix)
    if not present_months:
        return matrix

    canonical = [str(value).strip() for value in weekdays if str(value).strip()]
    if not canonical:
        return matrix

    for month in present_months:
        row_number = _to_int(month.get("row_start"))
        col_number = _to_int(month.get("column_start"))
        existing = month.get("weekdays", [])
        if row_number is None or col_number is None or not existing:
            continue

        header_row_number = row_number + 2
        _ensure_row(matrix, header_row_number)
        header_row = matrix[header_row_number - 1]
        for idx, current in enumerate(existing[:5]):
            _ensure_col(header_row, col_number + idx)
            next_value = canonical[idx] if idx < len(canonical) else str(current).strip()
            if next_value:
                header_row[col_number + idx - 1] = next_value

    return matrix


def _apply_holiday_updates_rows(rows: list[list[Any]], holidays: list[dict[str, Any]]) -> list[list[Any]]:
    cleaned = _normalize_holiday_updates(holidays)
    matrix = [list(HOLIDAY_HEADERS)]
    for item in cleaned:
        matrix.append(
            [
                item.get("name") or "",
                item.get("month_name") or "",
                item.get("day") if item.get("day") is not None else "",
                item.get("recurrence") or "",
                item.get("source") or "holidays",
                item.get("weekday") or "",
                item.get("year") or "",
                item.get("notes") or "",
            ]
        )
    return matrix if matrix else [[]]


def _apply_era_updates_rows(rows: list[list[Any]], era_events: list[dict[str, Any]]) -> list[list[Any]]:
    matrix = _pad_rows(rows)
    clear_rows, placements = _plan_era_event_placements(matrix, era_events)
    if not clear_rows and not placements:
        return matrix

    width = max((len(row) for row in matrix), default=0)
    for row_number in clear_rows:
        _ensure_row(matrix, row_number)
        row = matrix[row_number - 1]
        if len(row) < width:
            row.extend([""] * (width - len(row)))
        for col in range(width):
            row[col] = ""

    for row_number, col_number, text in placements:
        _ensure_row(matrix, row_number)
        row = matrix[row_number - 1]
        _ensure_col(row, col_number)
        row[col_number - 1] = text

    return matrix


def _normalize_holiday_updates(holidays: list[dict[str, Any]]) -> list[dict[str, Any]]:
    cleaned: list[dict[str, Any]] = []
    seen: set[tuple[str, str, int | None]] = set()

    for item in holidays:
        name = _as_text(item.get("name"))
        month_name = _as_text(item.get("month_name"))
        day = _to_int(item.get("day"))
        recurrence = _as_text(item.get("recurrence"))
        source = _as_text(item.get("source")) or "holidays"
        weekday = _as_text(item.get("weekday"))
        year = _as_text(item.get("year"))
        notes = _as_text(item.get("notes"))
        if not name:
            continue

        key = (name.lower(), month_name.lower(), day)
        if key in seen:
            continue
        seen.add(key)

        cleaned.append(
            {
                "name": name,
                "month_name": month_name or None,
                "day": day,
                "recurrence": recurrence or ("yearly" if month_name or day else None),
                "source": source,
                "weekday": weekday or None,
                "year": year or None,
                "notes": notes or None,
            }
        )

    cleaned.sort(
        key=lambda item: (
            str(item.get("month_name") or "").lower(),
            item.get("day") if item.get("day") is not None else 10**9,
            str(item.get("name") or "").lower(),
        )
    )
    return cleaned


def _sync_holidays_sheet_xlsx(workbook_path: Path, holidays: list[dict[str, Any]]) -> None:
    cleaned = _normalize_holiday_updates(holidays)
    if not workbook_path.exists():
        return

    wb = load_workbook(workbook_path)
    try:
        ws = wb["Holidays"] if "Holidays" in wb.sheetnames else wb.create_sheet("Holidays")
        existing_rows = [list(row) for row in ws.iter_rows(values_only=True)]
        existing_cleaned = _normalize_holiday_updates(_parse_holidays(existing_rows)) if existing_rows else []
        existing_headers = [str(ws.cell(1, col).value or "").strip() for col in range(1, len(HOLIDAY_HEADERS) + 1)]
        headers_match = all(
            _normalize_key(existing_headers[idx]) == _normalize_key(HOLIDAY_HEADERS[idx]) for idx in range(len(HOLIDAY_HEADERS))
        )

        if headers_match and existing_cleaned == cleaned:
            return

        _apply_holiday_updates_xlsx(ws, cleaned)
        wb.save(workbook_path)
    finally:
        wb.close()


def _plan_era_event_placements(
    rows: list[list[Any]],
    era_events: list[dict[str, Any]],
) -> tuple[list[int], list[tuple[int, int, str]]]:
    if not rows:
        return [], []

    max_cols = max((len(row) for row in rows), default=0)
    if max_cols == 0:
        return [], []

    block_start = 1  # zero-based row index of sheet row 2
    year_to_slots: dict[int, list[dict[str, Any]]] = {}
    clear_rows: set[int] = set()
    resolved_year_cells: dict[tuple[int, int], int] = {}

    while block_start < len(rows):
        year_row_idx = block_start + 1
        if year_row_idx >= len(rows):
            break

        event_row_numbers = []
        for offset in range(2, 6):
            idx = block_start + offset
            if idx < len(rows):
                event_row_numbers.append(idx + 1)
                clear_rows.add(idx + 1)

        year_row = rows[year_row_idx]
        previous_year: int | None = None
        year_row_number = year_row_idx + 1
        for col_idx in range(max_cols):
            raw_year = _cell_at(year_row, col_idx)
            year = _to_int(raw_year)
            if year is None:
                year = _infer_formula_number(raw_year, previous_year, resolved_year_cells)
            if year is None:
                continue
            previous_year = year
            resolved_year_cells[(year_row_number, col_idx + 1)] = year
            year_to_slots.setdefault(year, []).append(
                {
                    "column": col_idx + 1,
                    "event_rows": list(event_row_numbers),
                }
            )

        block_start += 7

    placements: list[tuple[int, int, str]] = []
    occupied: set[tuple[int, int]] = set()
    sortable = [item for item in era_events if _as_text(item.get("event"))]
    sortable.sort(
        key=lambda item: (
            _to_int(item.get("year")) if _to_int(item.get("year")) is not None else 10**12,
            _to_int(item.get("row_number")) if _to_int(item.get("row_number")) is not None else 10**12,
            _to_int(item.get("column")) if _to_int(item.get("column")) is not None else 10**12,
            _as_text(item.get("event")).lower(),
        )
    )

    for item in sortable:
        year = _to_int(item.get("year"))
        text = _as_text(item.get("event"))
        if year is None or not text:
            continue

        candidates = list(year_to_slots.get(year, []))
        if not candidates:
            continue

        preferred_col = _to_int(item.get("column"))
        preferred_row = _to_int(item.get("row_number"))

        if preferred_col is not None:
            candidates.sort(key=lambda slot: 0 if slot["column"] == preferred_col else 1)

        placed = False
        for slot in candidates:
            col_number = slot["column"]
            event_rows = slot["event_rows"]

            candidate_rows = list(event_rows)
            if preferred_row in event_rows:
                candidate_rows.sort(key=lambda row_number: 0 if row_number == preferred_row else 1)

            for row_number in candidate_rows:
                key = (row_number, col_number)
                if key in occupied:
                    continue
                placements.append((row_number, col_number, text))
                occupied.add(key)
                placed = True
                break

            if placed:
                break

    return sorted(clear_rows), placements


def _write_sheet_rows_google(client: SpreadsheetClient, sheet_name: str, rows: list[list[Any]]) -> None:
    worksheet = client.worksheet_by_title(sheet_name)
    worksheet.clear()
    if not rows:
        return
    worksheet.update(range_name="A1", values=rows)


def _detect_header_row_index(rows: list[list[Any]]) -> int:
    for idx, row in enumerate(rows[:10]):
        non_empty = sum(1 for cell in row if _as_text(cell))
        if non_empty >= 2:
            return idx
    return 0


def _calendar_header_indexes(headers: list[Any]) -> dict[str, int]:
    normalized = {re.sub(r"[^a-z0-9]+", "", _as_text(value).lower()): idx for idx, value in enumerate(headers)}
    return {
        "month_order": normalized.get("month", normalized.get("monthorder", 0)),
        "month_name": normalized.get("monthname", 1),
        "description": normalized.get("description", 2),
        "chore_name": normalized.get("chorename", 3),
        "chore_description": normalized.get("choredescription", 4),
        "deity_name": normalized.get("deityname", 5),
        "domain": normalized.get("domain", 6),
    }


def _column_from_group_key(key: Any) -> int | None:
    text = str(key or "").strip().lower()
    if not text.startswith("col_"):
        return None
    parsed = _to_int(text.split("_", 1)[1])
    if parsed is None or parsed < 1:
        return None
    return parsed


def _pad_rows(rows: list[list[Any]]) -> list[list[Any]]:
    if not rows:
        return []
    width = max((len(row) for row in rows), default=0)
    return [list(row) + [""] * (width - len(row)) for row in rows]


def _ensure_row(matrix: list[list[Any]], row_number: int) -> None:
    width = max((len(row) for row in matrix), default=0)
    while len(matrix) < row_number:
        matrix.append([""] * width)


def _ensure_col(row: list[Any], col_number: int) -> None:
    while len(row) < col_number:
        row.append("")


def _cell_at(row: list[Any], idx: int) -> Any:
    if idx < 0:
        return None
    if idx >= len(row):
        return None
    return row[idx]


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _as_text(value).lower())


def _to_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        parsed = float(text.replace(",", ""))
    except ValueError:
        return None
    if parsed.is_integer():
        return int(parsed)
    return None


def _infer_formula_number(
    value: Any,
    previous: int | None,
    resolved_cells: dict[tuple[int, int], int] | None = None,
) -> int | None:
    if previous is None:
        previous = None

    text = _as_text(value)
    if not text.startswith("="):
        return None

    ref_match = re.fullmatch(
        r"=\s*(?:(?:'[^']+'|[A-Za-z0-9_]+)!)?\$?([A-Z]+)\$?(\d+)(?:\s*([+-])\s*(\d+))?\s*",
        text,
        flags=re.IGNORECASE,
    )
    if ref_match:
        ref_col = column_index_from_string(ref_match.group(1).upper())
        ref_row = int(ref_match.group(2))
        base = None
        if resolved_cells is not None:
            base = resolved_cells.get((ref_row, ref_col))
        if base is None:
            base = previous
        if base is not None:
            sign = ref_match.group(3)
            amount = ref_match.group(4)
            if sign and amount:
                delta = int(amount) * (1 if sign == "+" else -1)
                return base + delta
            return base

    if previous is None:
        return None

    # Typical sheet formulas are incremental references like "=A3+1" or "=B4-1".
    delta_match = re.search(r"([+-])\s*(\d+)\s*$", text)
    if delta_match:
        sign = 1 if delta_match.group(1) == "+" else -1
        amount = int(delta_match.group(2))
        return previous + (sign * amount)

    ref_only = re.fullmatch(r"=\s*[\w.$!]+", text)
    if ref_only:
        return previous

    return None


def _clean_for_write(value: Any) -> Any:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _set_cell_if_writable(ws: Any, row_number: int, col_number: int, value: Any) -> None:
    cell = ws.cell(row_number, col_number)
    if isinstance(cell, MergedCell):
        return
    cell.value = value
