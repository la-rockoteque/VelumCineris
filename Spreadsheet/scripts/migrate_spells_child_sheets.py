from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path
from shutil import copy2
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


MODIFIER_HEADERS = [
    "Spell",
    "Spell Name",
    "Name",
    "Spell Row",
    "Modifier Index",
    "Modifier Type",
    "Modifier Subtype",
    "Modifier Dice Count",
    "Modifier Dice Type",
    "Modifier Fixed Value",
    "Modifier Use Primary Stat",
    "Modifier Duration",
    "Modifier Duration Unit",
    "Modifier Details",
    "Raw Modifier JSON",
    "Source",
]

SCALING_HEADERS = [
    "Spell",
    "Spell Name",
    "Name",
    "Spell Row",
    "Scaling Index",
    "Scaling",
    "Scaling Level",
    "Scaling Modifier",
    "Scaling Effect",
    "Scaling Dice Count",
    "Scaling Dice Type",
    "Scaling Fixed Value",
    "Source",
]

CONDITION_HEADERS = [
    "Spell",
    "Spell Name",
    "Name",
    "Spell Row",
    "Condition Index",
    "Condition",
    "Saving Throw",
    "Success",
    "Fail",
    "Skill Check",
    "Ability Check",
    "DC",
    "Trigger Text",
    "Source",
]


def as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).lower())


def parse_multi_values(value: Any) -> list[str]:
    text = as_text(value)
    if not text:
        return []
    parts = [item.strip() for item in re.split(r"[\n,;/|]+", text) if item.strip()]
    return parts if parts else [text]


def header_indexes(ws: Worksheet) -> dict[str, int]:
    idx: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = normalize_key(as_text(ws.cell(1, col).value))
        if key:
            idx[key] = col
    return idx


def find_col(cols: dict[str, int], *candidates: str) -> int | None:
    for candidate in candidates:
        normalized = normalize_key(candidate)
        if normalized in cols:
            return cols[normalized]
    return None


def read_cell(row: list[Any], col: int | None) -> Any:
    if col is None:
        return None
    index = col - 1
    if index < 0 or index >= len(row):
        return None
    return row[index]


def ensure_sheet(wb: Any, title: str, headers: list[str]) -> Worksheet:
    safe_title = title.translate(str.maketrans("", "", ':\\/?*[]'))
    existing = {normalize_key(name): name for name in wb.sheetnames}

    if normalize_key(title) in existing:
        ws = wb[existing[normalize_key(title)]]
    elif normalize_key(safe_title) in existing:
        ws = wb[existing[normalize_key(safe_title)]]
    else:
        ws = wb.create_sheet(title=safe_title)

    for i, header in enumerate(headers, start=1):
        ws.cell(1, i).value = header
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    return ws


def load_condition_terms(wb: Any) -> list[str]:
    if "Conditions" not in wb.sheetnames:
        return []
    ws = wb["Conditions"]
    terms: list[str] = []
    for row_num in range(2, ws.max_row + 1):
        value = as_text(ws.cell(row_num, 1).value)
        if not value:
            continue
        terms.append(value)
    # Longest-first reduces shorter-term overlaps.
    terms.sort(key=len, reverse=True)
    return terms


def detect_conditions(text: str, condition_terms: list[str]) -> list[str]:
    found: list[str] = []
    lower = text.lower()
    for term in condition_terms:
        probe = term.lower()
        pattern = r"\b" + re.escape(probe) + r"\b"
        if re.search(pattern, lower):
            found.append(term)
    deduped: list[str] = []
    seen: set[str] = set()
    for item in found:
        key = item.lower()
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)
    return deduped


def parse_modifiers_json(raw: str) -> list[dict[str, Any]]:
    text = raw.strip()
    if not text:
        return []
    try:
        payload = json.loads(text)
    except json.JSONDecodeError:
        return []
    if isinstance(payload, dict):
        payload = [payload]
    if not isinstance(payload, list):
        return []
    result: list[dict[str, Any]] = []
    for item in payload:
        if isinstance(item, dict):
            result.append(item)
    return result


def clear_inline_spell_columns(ws: Worksheet, row_num: int, cols_to_clear: list[int]) -> None:
    for col in cols_to_clear:
        ws.cell(row_num, col).value = None


def migrate(workbook_path: Path, backup: bool) -> dict[str, int]:
    if backup:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = workbook_path.with_name(f"{workbook_path.name}.bak.{stamp}")
        copy2(workbook_path, backup_path)

    wb = load_workbook(workbook_path)
    wb_values = load_workbook(workbook_path, data_only=True)
    try:
        if "Spells" not in wb.sheetnames:
            raise ValueError("Spells sheet not found in workbook.")
        spells_ws = wb["Spells"]
        values_spells_ws = wb_values["Spells"]
        cols = header_indexes(spells_ws)

        spell_name_col = find_col(cols, "Spell Name", "Name", "Spell")
        if spell_name_col is None:
            raise ValueError("Could not find Spell Name column in Spells sheet.")

        scaling_col = find_col(cols, "Scaling")
        scaling_level_col = find_col(cols, "Scaling Level")
        scaling_modifier_col = find_col(cols, "Scaling Modifier")
        scaling_effect_col = find_col(cols, "Scaling Effect")
        scaling_dice_count_col = find_col(cols, "Scaling Dice Count")
        scaling_dice_type_col = find_col(cols, "Scaling Dice Type")
        scaling_fixed_col = find_col(cols, "Fixed Value")

        modifier_details_col = find_col(cols, "Modifier Details")
        modifier_type_col = find_col(cols, "Modifier Type")
        modifier_subtype_col = find_col(cols, "Modifier Subtype")
        modifier_dice_count_col = find_col(cols, "Modifier Dice Count")
        modifier_dice_type_col = find_col(cols, "Modifier Dice Type")
        modifier_fixed_col = find_col(cols, "Modifier Fixed Value")
        modifier_primary_stat_col = find_col(cols, "Modifier Use Primary Stat")
        modifier_duration_col = find_col(cols, "Modifier Duration")
        modifier_duration_unit_col = find_col(cols, "Modifier Duration Unit")
        modifiers_json_col = find_col(cols, "Modifiers JSON")

        condition_col = find_col(cols, "Condition")
        saving_throw_col = find_col(cols, "Saving Throw")
        success_col = find_col(cols, "Success")
        fail_col = find_col(cols, "Fail")
        skill_check_col = find_col(cols, "Skill Check")
        ability_check_col = find_col(cols, "Ability Check")
        dc_col = find_col(cols, "DC")
        description_col = find_col(cols, "Description")

        modifiers_ws = ensure_sheet(wb, "Spells:Modifiers", MODIFIER_HEADERS)
        scaling_ws = ensure_sheet(wb, "Spells:Scaling", SCALING_HEADERS)
        conditions_ws = ensure_sheet(wb, "Spells:Conditions", CONDITION_HEADERS)

        condition_terms = load_condition_terms(wb)

        counts = {"modifiers": 0, "scaling": 0, "conditions": 0, "spells_touched": 0}
        modifiers_out: list[list[Any]] = []
        scaling_out: list[list[Any]] = []
        conditions_out: list[list[Any]] = []

        modifier_cols_to_clear = [
            col
            for col in [
                modifier_details_col,
                modifier_type_col,
                modifier_subtype_col,
                modifier_dice_count_col,
                modifier_dice_type_col,
                modifier_fixed_col,
                modifier_primary_stat_col,
                modifier_duration_col,
                modifier_duration_unit_col,
                modifiers_json_col,
            ]
            if col is not None
        ]
        scaling_cols_to_clear = [
            col
            for col in [
                scaling_col,
                scaling_level_col,
                scaling_modifier_col,
                scaling_effect_col,
                scaling_dice_count_col,
                scaling_dice_type_col,
                scaling_fixed_col,
            ]
            if col is not None
        ]
        condition_cols_to_clear = [
            col
            for col in [
                condition_col,
                saving_throw_col,
                success_col,
                fail_col,
                skill_check_col,
                ability_check_col,
                dc_col,
            ]
            if col is not None
        ]

        for row_num in range(2, spells_ws.max_row + 1):
            row = [spells_ws.cell(row_num, col).value for col in range(1, spells_ws.max_column + 1)]
            row_values = [values_spells_ws.cell(row_num, col).value for col in range(1, spells_ws.max_column + 1)]

            def v(col: int | None, *, fallback_to_formula: bool = True) -> Any:
                value_cell = read_cell(row_values, col)
                if value_cell not in (None, ""):
                    return value_cell
                if fallback_to_formula:
                    return read_cell(row, col)
                return None

            spell_name = as_text(v(spell_name_col))
            if not spell_name:
                continue

            touched = False

            raw_mod_json = as_text(v(modifiers_json_col))
            parsed_modifiers = parse_modifiers_json(raw_mod_json)

            for i, item in enumerate(parsed_modifiers, start=1):
                modifiers_out.append(
                    [
                        spell_name,
                        spell_name,
                        spell_name,
                        row_num,
                        i,
                        as_text(item.get("type")),
                        as_text(item.get("subtype")),
                        as_text(item.get("dice_count")),
                        as_text(item.get("dice_type")),
                        as_text(item.get("fixed_value")),
                        as_text(item.get("use_primary_stat") or item.get("usePrimaryStat")),
                        as_text(item.get("duration")),
                        as_text(item.get("duration_unit") or item.get("durationUnit")),
                        as_text(item.get("details")),
                        raw_mod_json,
                        "Modifiers JSON",
                    ]
                )
                counts["modifiers"] += 1
                touched = True

            inline_modifier_values = [
                as_text(v(modifier_details_col)),
                as_text(v(modifier_type_col)),
                as_text(v(modifier_subtype_col)),
                as_text(v(modifier_dice_count_col)),
                as_text(v(modifier_dice_type_col)),
                as_text(v(modifier_fixed_col)),
                as_text(v(modifier_primary_stat_col)),
                as_text(v(modifier_duration_col)),
                as_text(v(modifier_duration_unit_col)),
            ]
            if any(inline_modifier_values):
                details = inline_modifier_values[0]
                duplicate = False
                if details and parsed_modifiers:
                    duplicate = any(as_text(item.get("details")).strip().lower() == details.strip().lower() for item in parsed_modifiers)
                if not duplicate:
                    modifiers_out.append(
                        [
                            spell_name,
                            spell_name,
                            spell_name,
                            row_num,
                            len(parsed_modifiers) + 1,
                            inline_modifier_values[1],
                            inline_modifier_values[2],
                            inline_modifier_values[3],
                            inline_modifier_values[4],
                            inline_modifier_values[5],
                            inline_modifier_values[6],
                            inline_modifier_values[7],
                            inline_modifier_values[8],
                            inline_modifier_values[0],
                            raw_mod_json,
                            "Modifier columns",
                        ]
                    )
                    counts["modifiers"] += 1
                    touched = True

            scaling_values = [
                as_text(v(scaling_col)),
                as_text(v(scaling_level_col)),
                as_text(v(scaling_modifier_col)),
                as_text(v(scaling_effect_col)),
                as_text(v(scaling_dice_count_col)),
                as_text(v(scaling_dice_type_col)),
                as_text(v(scaling_fixed_col)),
            ]
            if any(scaling_values):
                scaling_out.append(
                    [
                        spell_name,
                        spell_name,
                        spell_name,
                        row_num,
                        1,
                        scaling_values[0],
                        scaling_values[1],
                        scaling_values[2],
                        scaling_values[3],
                        scaling_values[4],
                        scaling_values[5],
                        scaling_values[6],
                        "Scaling columns",
                    ]
                )
                counts["scaling"] += 1
                touched = True

            direct_condition_value = as_text(v(condition_col, fallback_to_formula=False))
            direct_conditions = parse_multi_values(direct_condition_value)
            condition_values = [item for item in direct_conditions if item]

            if not condition_values:
                scan_text = "\n".join(
                    [
                        as_text(v(description_col)),
                        as_text(v(modifier_details_col)),
                        as_text(v(scaling_col)),
                        as_text(v(success_col, fallback_to_formula=False)),
                        as_text(v(fail_col, fallback_to_formula=False)),
                    ]
                )
                condition_values = detect_conditions(scan_text, condition_terms)
                condition_source = "Detected from spell text"
            else:
                condition_source = "Condition column"

            if not condition_values:
                # Keep save/check metadata if present so data is not lost.
                if any(
                    as_text(v(col, fallback_to_formula=False))
                    for col in [saving_throw_col, success_col, fail_col, skill_check_col, ability_check_col, dc_col]
                    if col is not None
                ):
                    condition_values = [""]
                    condition_source = "Save/check metadata"

            for idx, condition_name in enumerate(condition_values, start=1):
                conditions_out.append(
                    [
                        spell_name,
                        spell_name,
                        spell_name,
                        row_num,
                        idx,
                        condition_name,
                        as_text(v(saving_throw_col, fallback_to_formula=False)),
                        as_text(v(success_col, fallback_to_formula=False)),
                        as_text(v(fail_col, fallback_to_formula=False)),
                        as_text(v(skill_check_col, fallback_to_formula=False)),
                        as_text(v(ability_check_col, fallback_to_formula=False)),
                        as_text(v(dc_col, fallback_to_formula=False)),
                        as_text(v(description_col)),
                        condition_source,
                    ]
                )
                counts["conditions"] += 1
                touched = True

            if touched:
                counts["spells_touched"] += 1
                clear_inline_spell_columns(spells_ws, row_num, modifier_cols_to_clear)
                clear_inline_spell_columns(spells_ws, row_num, scaling_cols_to_clear)
                clear_inline_spell_columns(spells_ws, row_num, condition_cols_to_clear)

        for row in modifiers_out:
            modifiers_ws.append(row)
        for row in scaling_out:
            scaling_ws.append(row)
        for row in conditions_out:
            conditions_ws.append(row)

        wb.save(workbook_path)
        return counts
    finally:
        wb_values.close()
        wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Normalize inline spell modifier/scaling/condition fields into child sheets.",
    )
    parser.add_argument(
        "--workbook",
        default="Spreadsheet/Orimond.xlsx",
        help="Path to workbook (default: Spreadsheet/Orimond.xlsx)",
    )
    parser.add_argument(
        "--no-backup",
        action="store_true",
        help="Skip timestamped backup copy before migration.",
    )
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    counts = migrate(workbook_path, backup=not args.no_backup)
    print(
        "Migration completed:",
        f"modifiers={counts['modifiers']}",
        f"scaling={counts['scaling']}",
        f"conditions={counts['conditions']}",
        f"spells_touched={counts['spells_touched']}",
    )


if __name__ == "__main__":
    main()
