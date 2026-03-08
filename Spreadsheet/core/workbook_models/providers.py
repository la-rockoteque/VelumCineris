from __future__ import annotations

from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from Spreadsheet.core.Helpers.sheets import SpreadsheetClient


class SheetProvider(ABC):
    @abstractmethod
    def sheet_names(self) -> list[str]:
        """Return all sheet names."""

    @abstractmethod
    def sheet_rows(self, sheet_name: str) -> list[list[Any]]:
        """Return raw sheet rows as a list of row lists."""


class XlsxSheetProvider(SheetProvider):
    def __init__(self, workbook_path: str | Path):
        self.workbook_path = Path(workbook_path)
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")
        self._cache: dict[str, list[list[Any]]] = {}
        self._sheet_names: list[str] | None = None

    def sheet_names(self) -> list[str]:
        if self._sheet_names is None:
            wb = load_workbook(self.workbook_path, read_only=True, data_only=True)
            self._sheet_names = list(wb.sheetnames)
            wb.close()
        return list(self._sheet_names)

    def sheet_rows(self, sheet_name: str) -> list[list[Any]]:
        if sheet_name in self._cache:
            return self._cache[sheet_name]

        wb = load_workbook(self.workbook_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()
        self._cache[sheet_name] = rows
        return rows


class GoogleSheetsProvider(SheetProvider):
    def __init__(self, spreadsheet_id: str, credentials_path: str | Path):
        self.client = SpreadsheetClient(
            spreadsheet_id,
            credentials_path=credentials_path,
            scopes=(
                "https://www.googleapis.com/auth/spreadsheets.readonly",
                "https://www.googleapis.com/auth/drive.readonly",
            ),
        )
        self._cache: dict[str, list[list[Any]]] = {}
        self._sheet_names: list[str] | None = None

    def sheet_names(self) -> list[str]:
        if self._sheet_names is None:
            self._sheet_names = self.client.list_sheet_names()
        return list(self._sheet_names)

    def sheet_rows(self, sheet_name: str) -> list[list[Any]]:
        if sheet_name in self._cache:
            return self._cache[sheet_name]

        rows = self.client.get_rows_by_title(sheet_name)
        self._cache[sheet_name] = rows
        return rows
