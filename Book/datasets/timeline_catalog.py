from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from Spreadsheet.datasets.content_registry import default_spreadsheet_id
from Spreadsheet.sheets import SpreadsheetClient

TIMELINE_SHEETS = {
    "era": "TimelineEra",
    "conflict": "TimelineConflict",
    "calendar": "TimelineCalendar",
    "naming": "TimelineNaming",
    "holidays": "TimelineHolidays",
    "present": "TimelinePresent",
}


def load_timeline_catalog(
    *,
    spreadsheet_id: str | None = None,
    credentials_path: str | None = None,
    xlsx_path: str | Path | None = None,
) -> dict[str, Any]:
    xlsx_candidate = _resolve_xlsx_path(xlsx_path)

    if xlsx_candidate is not None:
        sheets = _read_timeline_sheets_from_xlsx(xlsx_candidate)
    else:
        sheets = _read_timeline_sheets_from_google(
            spreadsheet_id=spreadsheet_id,
            credentials_path=credentials_path,
        )

    calendar_months = _parse_calendar_months(sheets["calendar"])
    naming_groups = _parse_naming_groups(sheets["naming"])
    present_months = _parse_present_months(sheets["present"], calendar_months)
    holidays = _merge_holidays(
        _parse_holidays(sheets["holidays"]),
        _extract_present_holidays(present_months),
    )

    return {
        "calendar_months": calendar_months,
        "naming_groups": naming_groups,
        "naming_template": _derive_naming_template(naming_groups),
        "weekdays": present_months[0]["weekdays"] if present_months else [],
        "holidays": holidays,
        "era_periods": _parse_era_periods(sheets["era"]),
        "era_events": _parse_era_events(sheets["era"]),
        "conflicts": _parse_conflicts(sheets["conflict"]),
        "present_months": present_months,
    }


def _resolve_xlsx_path(xlsx_path: str | Path | None) -> Path | None:
    candidates = [
        xlsx_path,
        os.getenv("VELUM_XLSX_PATH"),
        os.getenv("VELUM_TIMELINE_XLSX_PATH"),
        "Spreadsheet/Orimond.xlsx",
        "Timeline/Orimond Timeline.xlsx",
    ]
    for candidate in candidates:
        if not candidate:
            continue
        path = Path(candidate).expanduser()
        if path.exists():
            return path
    return None


def _read_timeline_sheets_from_xlsx(path: Path) -> dict[str, list[list[Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return {
            key: [list(row) for row in wb[name].iter_rows(values_only=True)]
            if name in wb.sheetnames
            else []
            for key, name in TIMELINE_SHEETS.items()
        }
    finally:
        wb.close()


def _read_timeline_sheets_from_google(
    *,
    spreadsheet_id: str | None,
    credentials_path: str | None,
) -> dict[str, list[list[Any]]]:
    client = SpreadsheetClient(
        spreadsheet_id or os.getenv("VELUM_TIMELINE_SPREADSHEET_ID") or default_spreadsheet_id("fantasy"),
        credentials_path=credentials_path or os.getenv("VELUM_GSHEETS_KEY_PATH"),
    )
    existing = set(client.list_sheet_names())
    out: dict[str, list[list[Any]]] = {}
    for key, name in TIMELINE_SHEETS.items():
        out[key] = client.get_rows_by_title(name) if name in existing else []
    return out


def _parse_calendar_months(rows: list[list[Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []

    header_row_idx = _detect_header_row_index(rows)
    headers = rows[header_row_idx] if header_row_idx < len(rows) else []
    header_map = _calendar_header_indexes(headers)
    months: list[dict[str, Any]] = []

    for idx in range(header_row_idx + 1, len(rows)):
        row = rows[idx]
        if not any(_as_text(cell) for cell in row):
            continue
        months.append(
            {
                "row_number": idx + 1,
                "month_order": _cell_at(row, header_map.get("month_order", 0)),
                "month_name": _cell_at(row, header_map.get("month_name", 1)),
                "description": _cell_at(row, header_map.get("description", 2)),
                "chore_name": _cell_at(row, header_map.get("chore_name", 3)),
                "chore_description": _cell_at(row, header_map.get("chore_description", 4)),
                "deity_name": _cell_at(row, header_map.get("deity_name", 5)),
                "domain": _cell_at(row, header_map.get("domain", 6)),
            }
        )
    return months


def _parse_naming_groups(rows: list[list[Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []

    headers = rows[0]
    seen_labels: dict[str, int] = {}
    groups: list[dict[str, Any]] = []

    for col_idx, raw_header in enumerate(headers, start=1):
        label = _as_text(raw_header)
        if not label:
            continue

        count = seen_labels.get(label, 0) + 1
        seen_labels[label] = count
        display = label if count == 1 else f"{label} ({count})"
        values = [
            value
            for row in rows[1:]
            if (value := _as_text(_cell_at(row, col_idx - 1)))
        ]
        groups.append({"key": f"col_{col_idx}", "label": display, "values": values})

    return groups


def _derive_naming_template(groups: list[dict[str, Any]]) -> str:
    if not groups:
        return "Year of the {A (Modifier)} {A (Phenomenon)} Century of the {A (Anchor)} {A (Modifier)}"

    def strip_variant(label: str) -> str:
        return re.sub(r"\s+\(\d+\)$", "", str(label or "")).strip() or "Part"

    primary = strip_variant(groups[0]["label"]) if len(groups) >= 1 else "A (Modifier)"
    phenomenon = strip_variant(groups[1]["label"]) if len(groups) >= 2 else "A (Phenomenon)"
    anchor = strip_variant(groups[2]["label"]) if len(groups) >= 3 else "A (Anchor)"
    trailing = strip_variant(groups[3]["label"]) if len(groups) >= 4 else "A (Modifier)"
    return f"Year of the {{{primary}}} {{{phenomenon}}} Century of the {{{anchor}}} {{{trailing}}}"


def _parse_holidays(rows: list[list[Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []

    header_map = _holiday_header_indexes(rows[0])
    if header_map.get("name") is not None:
        start_row = 2
    else:
        start_row = 1
        header_map = {}

    holidays: list[dict[str, Any]] = []
    for row_number in range(start_row, len(rows) + 1):
        row = rows[row_number - 1]
        if not any(_as_text(cell) for cell in row):
            continue

        name = _as_text(_cell_at(row, header_map.get("name", 0)))
        if not name:
            continue
        holidays.append(
            {
                "row_number": row_number,
                "name": name,
                "month_name": _as_text(_cell_at(row, header_map.get("month_name", 1))) or None,
                "day": _to_int(_cell_at(row, header_map.get("day", 2))),
                "recurrence": _as_text(_cell_at(row, header_map.get("recurrence", 3))) or None,
                "source": _as_text(_cell_at(row, header_map.get("source", 4))) or "holidays",
                "weekday": _as_text(_cell_at(row, header_map.get("weekday", 5))) or None,
                "year": _as_text(_cell_at(row, header_map.get("year", 6))) or None,
                "notes": _as_text(_cell_at(row, header_map.get("notes", 7))) or None,
            }
        )
    return holidays


def _parse_conflicts(rows: list[list[Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []

    conflicts: list[dict[str, Any]] = []
    for idx, row in enumerate(rows[1:], start=2):
        year = _to_int(_cell_at(row, 0))
        event = _as_text(_cell_at(row, 1))
        consequence = _as_text(_cell_at(row, 2))
        if year is None and not event and not consequence:
            continue
        conflicts.append(
            {
                "row_number": idx,
                "year": year,
                "event": event or None,
                "strategic_consequence": consequence or None,
            }
        )
    return conflicts


def _parse_era_periods(rows: list[list[Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []

    periods: list[dict[str, Any]] = []
    block_start = 1
    while block_start + 1 < len(rows):
        label_row = rows[block_start]
        year_row = rows[block_start + 1]
        width = max(len(label_row), len(year_row))
        current: dict[str, Any] | None = None
        previous_year: int | None = None

        for col_idx in range(1, width):
            raw_label = _as_text(_cell_at(label_row, col_idx))
            year = _to_int(_cell_at(year_row, col_idx))
            if year is None:
                year = _infer_formula_number(_cell_at(year_row, col_idx), previous_year)
            if year is None:
                continue
            previous_year = year
            era_name = _extract_era_name(raw_label)
            full_label = raw_label or era_name
            if current and current["era"] == era_name:
                current["end_year"] = year
                continue
            current = {
                "era": era_name or "Unspecified Era",
                "label": full_label or era_name or "Unspecified Era",
                "start_year": year,
                "end_year": year,
            }
            periods.append(current)

        block_start += 7

    return periods


def _parse_era_events(rows: list[list[Any]], limit: int = 1200) -> list[dict[str, Any]]:
    if not rows:
        return []

    events: list[dict[str, Any]] = []
    block_start = 1
    while block_start + 5 < len(rows):
        label_row = rows[block_start]
        year_row = rows[block_start + 1]
        event_rows = rows[block_start + 2 : block_start + 6]
        width = max(len(label_row), len(year_row), *(len(row) for row in event_rows), 0)
        previous_year: int | None = None

        for col_idx in range(1, width):
            year = _to_int(_cell_at(year_row, col_idx))
            if year is None:
                year = _infer_formula_number(_cell_at(year_row, col_idx), previous_year)
            if year is None:
                continue
            previous_year = year

            era_name = _extract_era_name(_as_text(_cell_at(label_row, col_idx)))
            for offset, event_row in enumerate(event_rows):
                event_type = _as_text(_cell_at(rows[block_start + 2 + offset], 0))
                event_text = _as_text(_cell_at(event_row, col_idx))
                if not event_text:
                    continue
                events.append(
                    {
                        "year": year,
                        "era": era_name or None,
                        "event_type": event_type or None,
                        "event": event_text,
                        "row_number": block_start + 3 + offset,
                        "column": col_idx + 1,
                    }
                )
                if len(events) >= limit:
                    return events

        block_start += 7

    return events


def _parse_present_months(
    rows: list[list[Any]],
    calendar_months: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    months: list[dict[str, Any]] = []
    if not rows:
        return months

    max_cols = max((len(row) for row in rows), default=0)
    row_idx = 0
    while row_idx + 2 < len(rows):
        year_row = rows[row_idx]
        month_row = rows[row_idx + 1]
        weekday_row = rows[row_idx + 2]

        group_found = False
        col = 0
        while col < max_cols:
            year_name = _as_text(_cell_at(year_row, col))
            month_name = _as_text(_cell_at(month_row, col))
            if month_name.startswith("=") and calendar_months:
                idx = len(months) % max(len(calendar_months), 1)
                month_name = _as_text(calendar_months[idx].get("month_name")).upper() or month_name

            weekdays = [_as_text(_cell_at(weekday_row, col + offset)) for offset in range(5)]
            weekdays = [value for value in weekdays if value]
            if month_name and len(weekdays) >= 3:
                group_found = True
                month_payload = {
                    "row_start": row_idx + 1,
                    "column_start": col + 1,
                    "year_name": year_name,
                    "month_name": month_name,
                    "weekdays": weekdays[:5],
                    "weeks": [],
                }

                last_day_number: int | None = None
                for week_idx in range(4):
                    number_row_idx = row_idx + 3 + (week_idx * 2)
                    events_row_idx = number_row_idx + 1
                    if number_row_idx >= len(rows):
                        break
                    number_row = rows[number_row_idx]
                    events_row = rows[events_row_idx] if events_row_idx < len(rows) else []

                    days: list[dict[str, Any]] = []
                    has_any = False
                    for day_offset in range(5):
                        day_col = col + day_offset
                        raw_day_number = _cell_at(number_row, day_col)
                        day_number = _to_int(raw_day_number)
                        if day_number is None:
                            day_number = _infer_formula_number(raw_day_number, last_day_number)
                        event_text = _as_text(_cell_at(events_row, day_col))
                        if day_number is not None or event_text:
                            has_any = True
                        if day_number is not None:
                            last_day_number = day_number
                        days.append(
                            {
                                "weekday": month_payload["weekdays"][day_offset]
                                if day_offset < len(month_payload["weekdays"])
                                else "",
                                "day": day_number,
                                "event": event_text or None,
                            }
                        )
                    if has_any:
                        month_payload["weeks"].append({"week_index": week_idx + 1, "days": days})

                months.append(month_payload)
                col += 6
                continue
            col += 1

        row_idx += 12 if group_found else 1

    return months


def _extract_present_holidays(present_months: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    seen: set[tuple[str, str, int | None]] = set()
    for month in present_months:
        month_name = _as_text(month.get("month_name"))
        for week in month.get("weeks", []):
            for day_payload in week.get("days", []):
                day_number = _to_int(day_payload.get("day"))
                event_text = _as_text(day_payload.get("event"))
                if not event_text:
                    continue
                for fragment in re.split(r"[\n/]+", event_text):
                    name = fragment.strip()
                    if not name:
                        continue
                    key = (name.lower(), month_name.lower(), day_number)
                    if key in seen:
                        continue
                    seen.add(key)
                    out.append(
                        {
                            "name": name,
                            "month_name": month_name or None,
                            "day": day_number,
                            "recurrence": "yearly",
                            "source": "present",
                            "weekday": _as_text(day_payload.get("weekday")) or None,
                            "year": _as_text(month.get("year_name")) or None,
                            "notes": None,
                        }
                    )
    return out


def _merge_holidays(base: list[dict[str, Any]], present: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: list[dict[str, Any]] = []
    seen: set[tuple[str, str, int | None]] = set()
    for item in base + present:
        name = _as_text(item.get("name"))
        month_name = _as_text(item.get("month_name"))
        day = _to_int(item.get("day"))
        if not name:
            continue
        key = (name.lower(), month_name.lower(), day)
        if key in seen:
            continue
        seen.add(key)
        payload = dict(item)
        payload["month_name"] = month_name or None
        payload["day"] = day
        merged.append(payload)
    merged.sort(
        key=lambda item: (
            str(item.get("month_name") or "").lower(),
            item.get("day") if item.get("day") is not None else 10**9,
            str(item.get("name") or "").lower(),
        )
    )
    return merged


def _extract_era_name(value: str) -> str:
    if not value:
        return ""
    parts = [part.strip() for part in value.splitlines() if part.strip()]
    if len(parts) >= 2:
        return parts[1]
    return parts[0] if parts else ""


def _detect_header_row_index(rows: list[list[Any]]) -> int:
    for idx, row in enumerate(rows[:3]):
        normalized = [re.sub(r"[^a-z]+", "", _as_text(cell).lower()) for cell in row if _as_text(cell)]
        if any("monthname" in value for value in normalized):
            return idx
    return 0


def _calendar_header_indexes(headers: list[Any]) -> dict[str, int]:
    normalized = {
        re.sub(r"[^a-z0-9]+", "", _as_text(value).lower()): idx
        for idx, value in enumerate(headers)
    }
    return {
        "month_order": normalized.get("month", normalized.get("monthorder", 0)),
        "month_name": normalized.get("monthname", 1),
        "description": normalized.get("description", 2),
        "chore_name": normalized.get("chorename", 3),
        "chore_description": normalized.get("choredescription", 4),
        "deity_name": normalized.get("deityname", 5),
        "domain": normalized.get("domain", 6),
    }


def _holiday_header_indexes(headers: list[Any]) -> dict[str, int]:
    normalized = {
        re.sub(r"[^a-z0-9]+", "", _as_text(value).lower()): idx
        for idx, value in enumerate(headers)
    }
    mapping: dict[str, int] = {}
    if "name" in normalized:
        mapping["name"] = normalized["name"]
    if "month" in normalized:
        mapping["month_name"] = normalized["month"]
    if "day" in normalized:
        mapping["day"] = normalized["day"]
    if "recurrence" in normalized:
        mapping["recurrence"] = normalized["recurrence"]
    if "source" in normalized:
        mapping["source"] = normalized["source"]
    if "weekday" in normalized:
        mapping["weekday"] = normalized["weekday"]
    if "year" in normalized:
        mapping["year"] = normalized["year"]
    if "notes" in normalized:
        mapping["notes"] = normalized["notes"]
    return mapping


def _cell_at(row: list[Any], idx: int) -> Any:
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _to_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    text = _as_text(value)
    match = re.search(r"-?\d+", text)
    return int(match.group(0)) if match else None


def _infer_formula_number(value: Any, previous_number: int | None) -> int | None:
    text = _as_text(value)
    if not text.startswith("=") or previous_number is None:
        return None
    if re.search(r"\+\s*1\b", text):
        return previous_number + 1
    return None
